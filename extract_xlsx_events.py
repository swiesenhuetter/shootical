from openpyxl import load_workbook
from collections import namedtuple
import datetime
import re
import locale
import html

EventData = namedtuple('EventData', ['date', 'time', 'descr'])


def xsl_get(local_name):
    worksheet = 'Hauptplan2023'
    wb = load_workbook(filename=local_name)
    xsl_cal = wb[worksheet]
    return xsl_cal


def event_time(hours_list):
    form = lambda s: re.sub(r'(\d\d)(\d\d)-(\d\d)(\d\d)', r'\1:\2 - \3:\4', s)
    hours_list = [form(hours) for hours in hours_list]
    return ' und '.join(hours_list)


def event_date(day):
    if not isinstance(day, datetime.datetime):
        return 'Error: {} is not date'.format(day)
    day_str = datetime.datetime.strftime(day, '%d.%m.%Y, %A')
    return day_str


def event_from_row(xsl_sheet, row_nr, hours_list):
    day = xsl_sheet.cell(row_nr, column=1).value
    event_name = xsl_sheet.cell(row=row_nr, column=6).value
    event_SL1 = xsl_sheet.cell(row=row_nr, column=4).value
    event_SL2 = xsl_sheet.cell(row=row_nr, column=5).value
    if (not event_SL1 and not event_SL2):
        event_desc = event_name
    else:
        event_desc = "{}   (SL: {} / {} )".format(event_name, event_SL1, event_SL2)
    evt_date = event_date(day)
    evt_time = event_time(hours_list)
    event_desc = html.escape(event_desc).encode('ascii', 'xmlcharrefreplace').decode('utf-8')
    evt = EventData(evt_date, evt_time, event_desc)
    return evt


def xl_to_events(xl_filename):
    locale.setlocale(locale.LC_ALL, 'german')
    xsl_cal = xsl_get(xl_filename)
    shtime = xsl_cal['C']
    events = []
    for t in shtime:
        match = re.findall(r'\d{4}-\d{4}', t.value)
        if match:
            evt = event_from_row(xsl_cal, t.row, match)
            events.append(evt)
    print("{} Events converted form {}".format(len(events), xl_filename))
    return events


def odd_month(date_str):
    match = re.findall(r'\d\d.[01][13579].\d{4}', date_str)
    if match:
        return True
    else:
        return False
