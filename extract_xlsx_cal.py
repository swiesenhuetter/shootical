from openpyxl import load_workbook
from icalendar import Calendar, Event
import datetime
import urllib.request as urq
import sys
import re

local_name = './data/PSUE_Schiesskalender2024.xlsx'
worksheet = 'Hauptplan2024'


# psue_url = 'http://www.psue.ch/calendar/PSUE_Schiesskalender2021.xlsx'


def xsl_get(url):
    wb = load_workbook(filename=local_name)
    xsl_cal = wb[worksheet]
    return xsl_cal


def event_datetime(day, hours_list):
    if not isinstance(day, datetime.datetime):
        print('{} {}'.format(day, hours_list))
        return
    start_txt = hours_list[0][0:4]
    end_txt = hours_list[-1][5:9]
    day_str = datetime.datetime.strftime(day, '%Y%m%d')
    start_ical_fmt = "{}T{}00".format(day_str, start_txt)
    end_ical_fmt = "{}T{}00".format(day_str, end_txt)
    return start_ical_fmt, end_ical_fmt


def event_from_row(xsl_sheet, row_nr, hours_list):
    day = xsl_sheet.cell(row_nr, column=1).value
    event_name = xsl_sheet.cell(row=row_nr, column=6).value
    event_SL1 = xsl_sheet.cell(row=row_nr, column=4).value
    event_SL2 = xsl_sheet.cell(row=row_nr, column=5).value
    if (not event_SL1 and not event_SL2):
        event_desc = event_name
    else:
        event_desc = "{}   (SL: {} / {} )".format(event_name, event_SL1, event_SL2)
    evt_from, evt_to = event_datetime(day, hours_list)
    evt_txt = "Von:{} Bis:{} Veranstaltung:{}".format(evt_from, evt_to, event_desc)
    cal_evt = Event()
    cal_evt['DTSTART'] = evt_from
    cal_evt['DTEND'] = evt_to
    cal_evt['SUMMARY'] = event_desc
    return cal_evt


def xl_to_calendar(xl_filename):
    xsl_cal = xsl_get(xl_filename)
    shtime = xsl_cal['C']
    cal = Calendar()
    for t in shtime:
        if t.value is None:
            continue
        match = re.findall(r'\d{4}-\d{4}', t.value)
        if match:
            evt = event_from_row(xsl_cal, t.row, match)
            cal.add_component(evt)
    cal_file_name = xl_filename.rsplit('.', 1)[0] + '.ics'
    cal_file = open(cal_file_name, 'wb')
    cal_file.write(cal.to_ical())
    print("{} Events converted form {}".format(len(cal.subcomponents), xl_filename))


def main():
    global local_name
    if len(sys.argv) >= 2:  # cmd line argument
        local_name = sys.argv[1]
    # else:    # download
    #     # urq.urlretrieve(psue_url, local_name)

    xl_to_calendar(local_name)


if __name__ == "__main__":
    main()
